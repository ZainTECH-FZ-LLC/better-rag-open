"""XLSX parser — openpyxl for sheets, data regions, formulas, and charts."""

from __future__ import annotations

import asyncio
import io

import structlog

from src.processing.parsers.base import DocumentParser, ParsedDocument

logger = structlog.get_logger()


class XLSXParser(DocumentParser):
    async def parse(self, file_bytes: bytes, filename: str) -> ParsedDocument:
        return await asyncio.to_thread(self._parse_sync, file_bytes, filename)

    def _parse_sync(self, file_bytes: bytes, filename: str) -> ParsedDocument:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

        sheets = []
        text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = {
                "name": sheet_name,
                "headers": [],
                "rows": [],
                "data_region": {"min_row": 1, "max_row": 0, "min_col": 1, "max_col": 0},
            }

            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(c.strip() for c in cells):
                    rows.append(cells)

            if rows:
                sheet_data["headers"] = rows[0]
                sheet_data["rows"] = rows
                sheet_data["data_region"]["max_row"] = len(rows)
                sheet_data["data_region"]["max_col"] = max(len(r) for r in rows)

            sheets.append(sheet_data)

            # Build text representation
            sheet_text = f"--- Sheet: {sheet_name} ---\n"
            if rows:
                # Header row
                sheet_text += " | ".join(rows[0]) + "\n"
                sheet_text += "-" * 40 + "\n"
                for row in rows[1:50]:  # Limit to first 50 rows for text
                    sheet_text += " | ".join(row) + "\n"
                if len(rows) > 50:
                    sheet_text += f"... ({len(rows) - 50} more rows)\n"
            text_parts.append(sheet_text)

        full_text = "\n\n".join(text_parts)
        wb.close()

        logger.info(
            "xlsx_parser.parsed",
            filename=filename,
            sheets=len(sheets),
        )

        return ParsedDocument(
            text=full_text,
            sheets=sheets,
            page_count=len(sheets),
            word_count=len(full_text.split()),
        )
