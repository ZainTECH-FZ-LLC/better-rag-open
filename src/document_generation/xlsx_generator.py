"""XLSX generator — creates Excel spreadsheets from structured specs."""

from __future__ import annotations

import asyncio
import uuid
from pathlib import Path

import structlog
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.document_generation.base import DocumentGenerator, DocumentSpec, GeneratedDocument

logger = structlog.get_logger()

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class XLSXGenerator(DocumentGenerator):
    """
    Generates Excel spreadsheets with data tables, formatting, and charts.

    Section types:
    - sheet: A worksheet with headers and data rows
    - chart_sheet: A worksheet with an embedded chart
    """

    async def generate(self, spec: DocumentSpec, output_dir: Path) -> GeneratedDocument:
        return await asyncio.to_thread(self._generate_sync, spec, output_dir)

    def _generate_sync(self, spec: DocumentSpec, output_dir: Path) -> GeneratedDocument:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for section in spec.sections:
            section_type = section.get("type", "sheet")

            if section_type == "sheet":
                self._add_data_sheet(wb, section)
            elif section_type == "chart_sheet":
                self._add_chart_sheet(wb, section)

        # If no sections were provided, create a default sheet
        if not wb.sheetnames:
            ws = wb.create_sheet("Sheet1")
            ws["A1"] = spec.title

        # Save
        filename = f"{_sanitize(spec.title)}_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = output_dir / filename
        wb.save(str(filepath))

        size = filepath.stat().st_size
        logger.info(
            "xlsx_generator.saved",
            filename=filename,
            size=size,
            sheets=len(wb.sheetnames),
        )

        return GeneratedDocument(
            filename=filename,
            filepath=filepath,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_type="xlsx",
            size_bytes=size,
            metadata={"sheet_count": len(wb.sheetnames)},
        )

    def _add_data_sheet(self, wb: Workbook, section: dict) -> None:
        """Add a formatted data sheet."""
        sheet_name = section.get("title", "Data")[:31]  # Excel limit
        ws = wb.create_sheet(sheet_name)

        headers = section.get("headers", [])
        rows_data = section.get("rows", [])

        if not headers:
            return

        # Header row
        for j, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=j, value=str(header))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Data rows
        for i, row in enumerate(rows_data, 2):
            for j, value in enumerate(row, 1):
                if j <= len(headers):
                    cell = ws.cell(row=i, column=j)
                    cell.border = THIN_BORDER

                    # Try to set numeric values
                    try:
                        cell.value = float(value)
                    except (ValueError, TypeError):
                        cell.value = str(value)

                    # Alternate row shading
                    if i % 2 == 0:
                        cell.fill = ALT_ROW_FILL

        # Auto-fit column widths (approximate)
        for j in range(1, len(headers) + 1):
            max_length = len(str(headers[j - 1]))
            for row in rows_data:
                if j - 1 < len(row):
                    max_length = max(max_length, len(str(row[j - 1])))
            ws.column_dimensions[get_column_letter(j)].width = min(max_length + 4, 30)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        if headers:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows_data) + 1}"

        # Formulas row (if specified)
        formulas = section.get("formulas", {})
        if formulas:
            formula_row = len(rows_data) + 2
            for col_letter, formula in formulas.items():
                cell = ws.cell(
                    row=formula_row,
                    column=_col_to_num(col_letter),
                    value=formula,
                )
                cell.font = Font(bold=True)

    def _add_chart_sheet(self, wb: Workbook, section: dict) -> None:
        """Add a sheet with an embedded chart."""
        # First add the data
        self._add_data_sheet(wb, section)

        headers = section.get("headers", [])
        rows_data = section.get("rows", [])
        sheet_name = section.get("title", "Data")[:31]
        ws = wb[sheet_name]

        chart_spec = section.get("chart", {})
        chart_type = chart_spec.get("type", "bar")
        chart_title = chart_spec.get("title", section.get("title", ""))

        if not headers or not rows_data:
            return

        # Create chart
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()

        chart.title = chart_title
        chart.width = 20
        chart.height = 12

        # Data reference (columns 2+ as series, column 1 as categories)
        data_ref = Reference(
            ws,
            min_col=2,
            min_row=1,
            max_col=len(headers),
            max_row=len(rows_data) + 1,
        )
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(rows_data) + 1)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, f"A{len(rows_data) + 4}")


def _sanitize(text: str) -> str:
    return "".join(c if c.isalnum() or c in " -_" else "" for c in text).strip()[:50]


def _col_to_num(col: str) -> int:
    """Convert Excel column letter to number (A=1, B=2, ...)."""
    result = 0
    for c in col.upper():
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result
